import difflib
import html
from abc import ABC, abstractmethod
from pathlib import Path
from typing import List, Tuple

import openpyxl
from pypdf import PdfReader

from models import DiffLine, DiffLineType, DocumentContent, FileType


class ContentExtractor(ABC):
    """ファイルからテキスト行を抽出する処理の共通インターフェース"""

    @abstractmethod
    def extract(self, path: Path) -> DocumentContent:
        raise NotImplementedError


class ExcelContentExtractor(ContentExtractor):
    """Excelファイル(.xlsx等)からシートごとの行内容を抽出する"""

    _CELL_DELIMITER: str = " | "

    def extract(self, path: Path) -> DocumentContent:
        workbook = openpyxl.load_workbook(path, data_only=True)
        lines: List[str] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            for row in sheet.iter_rows():
                cell_values = [self._cell_to_text(cell.value) for cell in row]
                lines.append(self._CELL_DELIMITER.join(cell_values))
        return DocumentContent(file_path=path, file_type=FileType.EXCEL, lines=tuple(lines))

    def _cell_to_text(self, value: object) -> str:
        return "" if value is None else str(value)


class PdfContentExtractor(ContentExtractor):
    """PDFファイルからページごとのテキスト行を抽出する"""

    def extract(self, path: Path) -> DocumentContent:
        reader = PdfReader(str(path))
        lines: List[str] = []
        for page_number, page in enumerate(reader.pages, start=1):
            lines.append(f"=== Page {page_number} ===")
            page_text = page.extract_text() or ""
            lines.extend(page_text.splitlines())
        return DocumentContent(file_path=path, file_type=FileType.PDF, lines=tuple(lines))


class ContentExtractorFactory:
    """ファイル拡張子に応じた ContentExtractor を提供する"""

    _EXCEL_SUFFIXES: Tuple[str, ...] = (".xlsx", ".xlsm")
    _PDF_SUFFIXES: Tuple[str, ...] = (".pdf",)

    def create_for(self, path: Path) -> ContentExtractor:
        suffix = path.suffix.lower()
        if suffix in self._EXCEL_SUFFIXES:
            return ExcelContentExtractor()
        if suffix in self._PDF_SUFFIXES:
            return PdfContentExtractor()
        raise ValueError(f"未対応のファイル形式です: {suffix}")


class DiffService:
    """2つの DocumentContent の行内容を比較し、差分行のリストを生成する"""

    def compute(self, left: DocumentContent, right: DocumentContent) -> List[DiffLine]:
        matcher = difflib.SequenceMatcher(None, left.lines, right.lines)
        diff_lines: List[DiffLine] = []
        for tag, left_start, left_end, right_start, right_end in matcher.get_opcodes():
            if tag == "equal":
                diff_lines.extend(
                    self._build_equal_lines(left, left_start, left_end, right, right_start)
                )
            elif tag == "replace":
                diff_lines.extend(
                    self._build_replace_lines(
                        left, left_start, left_end, right, right_start, right_end
                    )
                )
            elif tag == "delete":
                diff_lines.extend(self._build_delete_lines(left, left_start, left_end))
            elif tag == "insert":
                diff_lines.extend(self._build_insert_lines(right, right_start, right_end))
        return diff_lines

    def _build_equal_lines(
        self,
        left: DocumentContent,
        left_start: int,
        left_end: int,
        right: DocumentContent,
        right_start: int,
    ) -> List[DiffLine]:
        result: List[DiffLine] = []
        for offset, left_index in enumerate(range(left_start, left_end)):
            right_index = right_start + offset
            result.append(
                DiffLine(
                    line_type=DiffLineType.EQUAL,
                    left_line_no=left_index + 1,
                    left_text=left.lines[left_index],
                    right_line_no=right_index + 1,
                    right_text=right.lines[right_index],
                )
            )
        return result

    def _build_replace_lines(
        self,
        left: DocumentContent,
        left_start: int,
        left_end: int,
        right: DocumentContent,
        right_start: int,
        right_end: int,
    ) -> List[DiffLine]:
        result: List[DiffLine] = []
        left_count = left_end - left_start
        right_count = right_end - right_start
        paired_count = min(left_count, right_count)

        for offset in range(paired_count):
            left_index = left_start + offset
            right_index = right_start + offset
            result.append(
                DiffLine(
                    line_type=DiffLineType.REPLACE,
                    left_line_no=left_index + 1,
                    left_text=left.lines[left_index],
                    right_line_no=right_index + 1,
                    right_text=right.lines[right_index],
                )
            )

        if left_count > paired_count:
            result.extend(
                self._build_delete_lines(left, left_start + paired_count, left_end)
            )
        if right_count > paired_count:
            result.extend(
                self._build_insert_lines(right, right_start + paired_count, right_end)
            )
        return result

    def _build_delete_lines(
        self, left: DocumentContent, left_start: int, left_end: int
    ) -> List[DiffLine]:
        return [
            DiffLine(
                line_type=DiffLineType.DELETE,
                left_line_no=left_index + 1,
                left_text=left.lines[left_index],
                right_line_no=None,
                right_text=None,
            )
            for left_index in range(left_start, left_end)
        ]

    def _build_insert_lines(
        self, right: DocumentContent, right_start: int, right_end: int
    ) -> List[DiffLine]:
        return [
            DiffLine(
                line_type=DiffLineType.INSERT,
                left_line_no=None,
                left_text=None,
                right_line_no=right_index + 1,
                right_text=right.lines[right_index],
            )
            for right_index in range(right_start, right_end)
        ]


class HtmlDiffReportRenderer:
    """差分行のリストから左右対比のHTMLレポートを生成する"""

    _ROW_COLORS = {
        DiffLineType.EQUAL: "#ffffff",
        DiffLineType.INSERT: "#d4f8d4",
        DiffLineType.DELETE: "#f8d4d4",
        DiffLineType.REPLACE: "#fdf3c4",
    }

    def render(
        self, left: DocumentContent, right: DocumentContent, diff_lines: List[DiffLine]
    ) -> str:
        rows_html = "\n".join(self._render_row(diff_line) for diff_line in diff_lines)
        return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>差分比較レポート</title>
<style>
  body {{ font-family: "Segoe UI", sans-serif; margin: 16px; }}
  table {{ border-collapse: collapse; width: 100%; table-layout: fixed; }}
  th, td {{ border: 1px solid #ccc; padding: 4px 8px; vertical-align: top; word-break: break-all; }}
  th {{ background: #333; color: #fff; position: sticky; top: 0; }}
  .line-no {{ width: 48px; color: #888; text-align: right; }}
  .text {{ white-space: pre-wrap; font-family: Consolas, monospace; }}
  h1 {{ font-size: 18px; }}
</style>
</head>
<body>
<h1>差分比較レポート</h1>
<p>左: {html.escape(left.file_name)} / 右: {html.escape(right.file_name)}</p>
<table>
<thead>
<tr>
<th class="line-no">#</th><th>{html.escape(left.file_name)}</th>
<th class="line-no">#</th><th>{html.escape(right.file_name)}</th>
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
</body>
</html>
"""

    def _render_row(self, diff_line: DiffLine) -> str:
        color = self._ROW_COLORS[diff_line.line_type]
        left_no = "" if diff_line.left_line_no is None else str(diff_line.left_line_no)
        right_no = "" if diff_line.right_line_no is None else str(diff_line.right_line_no)
        left_text = html.escape(diff_line.left_text) if diff_line.left_text is not None else ""
        right_text = html.escape(diff_line.right_text) if diff_line.right_text is not None else ""
        return (
            f'<tr style="background-color: {color};">'
            f'<td class="line-no">{left_no}</td><td class="text">{left_text}</td>'
            f'<td class="line-no">{right_no}</td><td class="text">{right_text}</td>'
            f"</tr>"
        )


class ReportFileWriter:
    """生成したHTML文字列をファイルに書き出す"""

    def write(self, html_content: str, output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(html_content, encoding="utf-8")
