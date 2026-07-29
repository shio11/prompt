import csv
import re
from abc import ABC, abstractmethod
from pathlib import Path
from typing import List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook

from models import CriterionResult, EvaluationResult, Prompt


class PromptFileFormatError(Exception):
    """未対応のファイル形式、または必要な列が見つからない場合に送出する例外。"""


class ScoringCriterion(ABC):
    """採点基準の共通インターフェース。具象クラスは `evaluate` のみを実装する。"""

    def __init__(self, weight: float) -> None:
        if weight <= 0:
            raise ValueError("weight は正の値でなければなりません")
        self._weight = weight

    @property
    def weight(self) -> float:
        return self._weight

    @property
    @abstractmethod
    def name(self) -> str: ...

    @abstractmethod
    def evaluate(self, prompt: Prompt) -> CriterionResult: ...

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}(weight={self._weight})"


class TargetSpecificationCriterion(ScoringCriterion):
    """対象のシート・セル範囲・テーブルなどが明示されているかを判定する。"""

    _KEYWORDS = ("シート", "セル", "範囲", "テーブル", "列", "行", "表")
    _CELL_REF_PATTERN = re.compile(r"[A-Za-z]{1,3}\d+(:[A-Za-z]{1,3}\d+)?")

    def __init__(self, weight: float = 20.0) -> None:
        super().__init__(weight)

    @property
    def name(self) -> str:
        return "対象範囲の明示"

    def evaluate(self, prompt: Prompt) -> CriterionResult:
        matched = any(kw in prompt.text for kw in self._KEYWORDS) or bool(
            self._CELL_REF_PATTERN.search(prompt.text)
        )
        if matched:
            return CriterionResult(self.name, self.weight, self.weight, True, "対象範囲が明示されています。")
        return CriterionResult(
            self.name,
            self.weight,
            0.0,
            False,
            "対象のシート・セル範囲・テーブル名などが明示されていません。「Sheet1のA1:D10」のように具体的に指定してください。",
        )


class ActionVerbCriterion(ScoringCriterion):
    """実行してほしい操作が具体的な動詞で示されているかを判定する。"""

    _VERBS = (
        "追加", "削除", "更新", "計算", "集計", "抽出", "並び替え", "ソート",
        "フィルタ", "作成", "変換", "結合", "分割", "置換", "書式設定",
        "グラフ化", "入力", "コピー", "移動",
    )

    def __init__(self, weight: float = 20.0) -> None:
        super().__init__(weight)

    @property
    def name(self) -> str:
        return "操作内容の具体性"

    def evaluate(self, prompt: Prompt) -> CriterionResult:
        matched = any(verb in prompt.text for verb in self._VERBS)
        if matched:
            return CriterionResult(self.name, self.weight, self.weight, True, "実行する操作が具体的に示されています。")
        return CriterionResult(
            self.name,
            self.weight,
            0.0,
            False,
            "「追加する」「集計する」のように、実行してほしい具体的な操作を動詞で示してください。",
        )


class OutputFormatCriterion(ScoringCriterion):
    """期待する出力形式が明示されているかを判定する。"""

    _KEYWORDS = ("出力して", "表示して", "返して", "一覧で", "グラフで", "表形式で", "まとめて", "作成して")

    def __init__(self, weight: float = 15.0) -> None:
        super().__init__(weight)

    @property
    def name(self) -> str:
        return "出力形式の明示"

    def evaluate(self, prompt: Prompt) -> CriterionResult:
        matched = any(kw in prompt.text for kw in self._KEYWORDS)
        if matched:
            return CriterionResult(self.name, self.weight, self.weight, True, "期待する出力形式が明示されています。")
        return CriterionResult(
            self.name,
            self.weight,
            0.0,
            False,
            "結果をどう受け取りたいか（表で出力／グラフを作成 等）を明示すると、意図しない出力を防げます。",
        )


class AmbiguousExpressionCriterion(ScoringCriterion):
    """「適当に」など曖昧な表現が含まれていないかを判定する。"""

    _AMBIGUOUS_WORDS = ("適当に", "いい感じに", "なるべく", "できれば", "多分", "たぶん", "うまく", "よしなに")

    def __init__(self, weight: float = 20.0) -> None:
        super().__init__(weight)

    @property
    def name(self) -> str:
        return "曖昧な表現の排除"

    def evaluate(self, prompt: Prompt) -> CriterionResult:
        found = [word for word in self._AMBIGUOUS_WORDS if word in prompt.text]
        if not found:
            return CriterionResult(self.name, self.weight, self.weight, True, "曖昧な表現は見つかりませんでした。")
        return CriterionResult(
            self.name,
            self.weight,
            0.0,
            False,
            f"曖昧な表現が含まれています（{'、'.join(found)}）。具体的な条件・数値・基準に置き換えてください。",
        )


class ConditionCriterion(ScoringCriterion):
    """条件分岐や例外・除外ルールが必要な場合に明示されているかを判定する。"""

    _KEYWORDS = ("もし", "場合", "条件", "以上", "以下", "未満", "除く", "を除いて", "なら")

    def __init__(self, weight: float = 15.0) -> None:
        super().__init__(weight)

    @property
    def name(self) -> str:
        return "条件・制約の明示"

    def evaluate(self, prompt: Prompt) -> CriterionResult:
        matched = any(kw in prompt.text for kw in self._KEYWORDS)
        if matched:
            return CriterionResult(self.name, self.weight, self.weight, True, "条件・制約が明示されています。")
        return CriterionResult(
            self.name,
            self.weight,
            self.weight * 0.5,
            True,
            "条件分岐や除外ルールが必要な場合は「〇〇の場合は除く」のように明示すると精度が上がります（該当しない場合は無視して構いません）。",
        )


class LengthSpecificityCriterion(ScoringCriterion):
    """プロンプトの文字数から、情報量が十分かを判定する。"""

    _MIN_LENGTH = 15
    _IDEAL_MIN_LENGTH = 30

    def __init__(self, weight: float = 10.0) -> None:
        super().__init__(weight)

    @property
    def name(self) -> str:
        return "情報量"

    def evaluate(self, prompt: Prompt) -> CriterionResult:
        length = len(prompt.text.strip())
        if length < self._MIN_LENGTH:
            return CriterionResult(
                self.name, self.weight, 0.0, False,
                "プロンプトが短すぎます。対象・操作・条件・出力形式を具体的に追記してください。",
            )
        if length < self._IDEAL_MIN_LENGTH:
            return CriterionResult(
                self.name, self.weight, self.weight * 0.5, False,
                "もう少し詳細な情報（対象範囲や条件など）を追加すると精度が上がります。",
            )
        return CriterionResult(self.name, self.weight, self.weight, True, "十分な情報量があります。")


class PromptScorer:
    """複数のScoringCriterionを組み合わせてプロンプトを採点するサービス。"""

    def __init__(self, criteria: Optional[Sequence[ScoringCriterion]] = None) -> None:
        self._criteria: Tuple[ScoringCriterion, ...] = tuple(criteria) if criteria else self._default_criteria()

    @staticmethod
    def _default_criteria() -> Tuple[ScoringCriterion, ...]:
        return (
            TargetSpecificationCriterion(),
            ActionVerbCriterion(),
            OutputFormatCriterion(),
            AmbiguousExpressionCriterion(),
            ConditionCriterion(),
            LengthSpecificityCriterion(),
        )

    @property
    def criteria(self) -> Tuple[ScoringCriterion, ...]:
        return self._criteria

    def score(self, prompt: Prompt) -> EvaluationResult:
        results = tuple(criterion.evaluate(prompt) for criterion in self._criteria)
        return EvaluationResult(prompt=prompt, criterion_results=results)

    def __repr__(self) -> str:
        return f"PromptScorer(criteria={[c.name for c in self._criteria]!r})"


class PromptRepository:
    """CSV/Excelファイルからプロンプト一覧を読み込むサービス。"""

    _SUPPORTED_EXCEL_SUFFIXES = (".xlsx", ".xlsm")

    def __init__(self, text_column: str = "prompt", id_column: Optional[str] = None) -> None:
        self._text_column = text_column
        self._id_column = id_column

    def load(self, path: Path) -> List[Prompt]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv(path)
        if suffix in self._SUPPORTED_EXCEL_SUFFIXES:
            return self._load_xlsx(path)
        raise PromptFileFormatError(f"未対応のファイル形式です: {suffix}")

    def _load_csv(self, path: Path) -> List[Prompt]:
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None or self._text_column not in reader.fieldnames:
                raise PromptFileFormatError(f"列 '{self._text_column}' が見つかりません。")
            prompts: List[Prompt] = []
            for index, row in enumerate(reader, start=1):
                text = (row.get(self._text_column) or "").strip()
                if not text:
                    continue
                prompt_id = self._resolve_id(row.get(self._id_column) if self._id_column else None, index)
                prompts.append(Prompt(id=prompt_id, text=text))
            return prompts

    def _load_xlsx(self, path: Path) -> List[Prompt]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                return []
            column_index = {str(name): idx for idx, name in enumerate(header) if name is not None}
            if self._text_column not in column_index:
                raise PromptFileFormatError(f"列 '{self._text_column}' が見つかりません。")
            text_idx = column_index[self._text_column]
            id_idx = column_index.get(self._id_column) if self._id_column else None

            prompts: List[Prompt] = []
            for index, row in enumerate(rows, start=1):
                raw_text = row[text_idx] if text_idx < len(row) else None
                text = str(raw_text).strip() if raw_text is not None else ""
                if not text:
                    continue
                raw_id = row[id_idx] if id_idx is not None and id_idx < len(row) else None
                prompt_id = self._resolve_id(raw_id, index)
                prompts.append(Prompt(id=prompt_id, text=text))
            return prompts
        finally:
            workbook.close()

    @staticmethod
    def _resolve_id(raw_id: object, fallback_index: int) -> str:
        if raw_id is None or str(raw_id).strip() == "":
            return str(fallback_index)
        return str(raw_id).strip()


class ScoreReportWriter:
    """採点結果をCSV/Excelファイルに出力するサービス。"""

    _SUPPORTED_EXCEL_SUFFIXES = (".xlsx", ".xlsm")

    def write(self, results: Sequence[EvaluationResult], path: Path) -> None:
        if not results:
            raise ValueError("results が空です")
        header, rows = self._build_rows(results)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            self._write_csv(header, rows, path)
        elif suffix in self._SUPPORTED_EXCEL_SUFFIXES:
            self._write_xlsx(header, rows, path)
        else:
            raise PromptFileFormatError(f"未対応のファイル形式です: {suffix}")

    @staticmethod
    def _build_rows(results: Sequence[EvaluationResult]) -> Tuple[List[str], List[List[object]]]:
        criterion_names = [result.name for result in results[0].criterion_results]
        header = ["id", "prompt", "total_score", "feedback"] + criterion_names
        rows: List[List[object]] = []
        for result in results:
            row: List[object] = [result.prompt.id, result.prompt.text, result.total_score, result.feedback]
            row += [f"{cr.score:g}/{cr.weight:g}" for cr in result.criterion_results]
            rows.append(row)
        return header, rows

    @staticmethod
    def _write_csv(header: List[str], rows: List[List[object]], path: Path) -> None:
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(rows)

    @staticmethod
    def _write_xlsx(header: List[str], rows: List[List[object]], path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(header)
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
