from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import quote

import msal
import requests
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import FileLinkInfo, SharePointConfig


class GraphAuthService:
    """Microsoft Graph APIのアクセストークン取得のみを担当する"""

    _SCOPES: List[str] = ["https://graph.microsoft.com/.default"]

    def __init__(self, config: SharePointConfig) -> None:
        self._app = msal.ConfidentialClientApplication(
            client_id=config.client_id,
            client_credential=config.client_secret,
            authority=f"https://login.microsoftonline.com/{config.tenant_id}",
        )

    def get_access_token(self) -> str:
        result = self._app.acquire_token_for_client(scopes=self._SCOPES)
        if "access_token" not in result:
            error = result.get("error_description", "認証に失敗しました")
            raise RuntimeError(f"アクセストークンの取得に失敗しました: {error}")
        return str(result["access_token"])


class GraphApiClient:
    """Microsoft Graph APIへのHTTP通信のみを担当する"""

    _BASE_URL: str = "https://graph.microsoft.com/v1.0"

    def __init__(self, auth_service: GraphAuthService) -> None:
        self._auth_service = auth_service

    def get(self, path: str) -> Dict[str, Any]:
        url = path if path.startswith("http") else f"{self._BASE_URL}{path}"
        response = requests.get(url, headers=self._headers())
        response.raise_for_status()
        return response.json()

    def get_all_values(self, path: str) -> List[Dict[str, Any]]:
        values: List[Dict[str, Any]] = []
        next_url: Optional[str] = path
        while next_url is not None:
            payload = self.get(next_url)
            values.extend(payload.get("value", []))
            next_url = payload.get("@odata.nextLink")
        return values

    def _headers(self) -> Dict[str, str]:
        return {"Authorization": f"Bearer {self._auth_service.get_access_token()}"}


class SharePointSiteResolver:
    """サイトIDとドキュメントライブラリ(ドライブ)IDの解決のみを担当する"""

    def __init__(self, api_client: GraphApiClient) -> None:
        self._api_client = api_client

    def resolve_drive_id(self, hostname: str, site_relative_path: str) -> str:
        encoded_path = quote(site_relative_path, safe="/")
        site = self._api_client.get(f"/sites/{hostname}:{encoded_path}")
        drive = self._api_client.get(f"/sites/{site['id']}/drive")
        return str(drive["id"])


class SharePointFolderFileRepository:
    """指定フォルダ配下のファイル一覧(リンク含む)の取得のみを担当する"""

    def __init__(self, api_client: GraphApiClient) -> None:
        self._api_client = api_client

    def find_all_files(self, drive_id: str, folder_path: str) -> List[FileLinkInfo]:
        encoded_path = quote(folder_path, safe="/")
        root_children_url = f"/drives/{drive_id}/root:/{encoded_path}:/children"
        return self._collect_files(drive_id, root_children_url, folder_path)

    def _collect_files(
        self, drive_id: str, children_url: str, current_folder_path: str
    ) -> List[FileLinkInfo]:
        files: List[FileLinkInfo] = []
        for item in self._api_client.get_all_values(children_url):
            if "folder" in item:
                sub_folder_path = f"{current_folder_path}/{item['name']}"
                sub_children_url = f"/drives/{drive_id}/items/{item['id']}/children"
                files.extend(self._collect_files(drive_id, sub_children_url, sub_folder_path))
            elif "file" in item:
                files.append(
                    FileLinkInfo(
                        file_name=item["name"],
                        folder_path=current_folder_path,
                        web_url=item["webUrl"],
                        size_bytes=item.get("size", 0),
                    )
                )
        return files


class FileLinkExcelExporter:
    """ファイルリンク一覧のExcel出力のみを担当する"""

    _HEADERS: List[str] = ["ファイル名", "フォルダパス", "リンクURL", "サイズ(バイト)"]

    def export(self, file_links: List[FileLinkInfo], output_path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "SharePointリンク一覧"
        self._write_header(worksheet)
        self._write_rows(worksheet, file_links)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _write_header(self, worksheet: Worksheet) -> None:
        worksheet.append(self._HEADERS)

    def _write_rows(self, worksheet: Worksheet, file_links: List[FileLinkInfo]) -> None:
        for row_index, file_link in enumerate(file_links, start=2):
            worksheet.cell(row=row_index, column=1, value=file_link.file_name)
            worksheet.cell(row=row_index, column=2, value=file_link.folder_path)
            link_cell = worksheet.cell(row=row_index, column=3, value=file_link.web_url)
            link_cell.hyperlink = file_link.web_url
            link_cell.style = "Hyperlink"
            worksheet.cell(row=row_index, column=4, value=file_link.size_bytes)
